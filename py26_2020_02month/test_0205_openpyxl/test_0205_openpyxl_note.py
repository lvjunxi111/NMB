#!/usr/bin/python3
# -*-coding:utf-8-*-
# @time      :2020/2/6 20:25
# @Author    :LvJunXi
# @Mail      :18912945952@163.com
# @File      :test_0205_openpyxl_note.py
# @Software  :PyCharm
# ----------------------------------
import openpyxl

# openpyxl 模块读取Excel.xlsx文件,返回一个workbook对象（工作簿）
# workbook  [工作簿]
# sheet     [表单]
# cell      [格子对象]

wb = openpyxl.load_workbook("cases.xlsx")
# print(wb)

# 选取工作簿中的表单
sheet = wb["register"]
# print(sh)

# 读取表单某个各自信息
data = sheet.cell(row=3,column=3).value
# print(data)

# 想excel写入数据
# sh.cell(row=10,column=10,value="写入内容")
# 保存文档,保存文件时文件不能是打开状态。
# wb.save("cases.xlsx")

# 按行获取表单中数据，每行数据放在元组中。
# datas = list(sheet.rows)
# print(li)
# li = []
# for i in datas[2]:
#     li.append(i.value)
"""
# 列表推导式
# datas = list(sheet.rows)
li = [i.value for i in datas[2]]
"""

# print(datas)
# 按列获取表单中数据，每列数据放在元组中。
datas_1 =sheet.columns
# print(list(datas_1))

# 获取最大行 sheet.max_row
# 获取最大列 sheet.max_column






# ---------作业-------------------
# 1、同步敲完老师上课代码（不用提交）
# 2、将前面上课login_check这个功能函数的单元测试代码，进行代码和代码数据分离处理，用例数据保存在列表中
# 3、通过代码读取附件文件中的register这个表单的第一行和第二行数据，想办法保存为一个字典：格式如下：
# {'case_id': 1, 'title': '正常注册',
# 'data': "('python1', '123456', '123456')",
# 'expected': '{"code": 1, "msg": "注册成功"}',
# 'result': None}
#